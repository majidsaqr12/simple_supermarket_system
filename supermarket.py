import random
from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook
# -*- coding: UTF-8 -*-


class supermarket:
    def __init__(self, root):
        self.root = root
        self.root.geometry('1250x690+60+2')
        self.root.title('Supermarket System')
        self.root.resizable(False, False)
        title = Label(self.root, text='SuperMarket System', fg='white', bg='#0B2F3A', font=('tajawal', 16))
        title.pack(fill=X)
        # ============ V Customer =========
        self.name = StringVar()
        self.phone = StringVar()
        self.bill = StringVar()
        x = random.randint(1, 100)
        self.bill.set(str(x))

        self.x = IntVar()

        # =========== V section 1 ========
        self.s1 = IntVar()
        self.s2 = IntVar()
        self.s3 = IntVar()
        self.s4 = IntVar()
        self.s5 = IntVar()
        self.s6 = IntVar()
        self.s7 = IntVar()
        self.s8 = IntVar()
        self.s9 = IntVar()
        self.s10 = IntVar()
        self.s11 = IntVar()
        self.s12 = IntVar()
        self.s13 = IntVar()
        self.s14 = IntVar()
        self.s15 = IntVar()

        # =========== V section 2 ========
        self.ss1 = IntVar()
        self.ss2 = IntVar()
        self.ss3 = IntVar()
        self.ss4 = IntVar()
        self.ss5 = IntVar()
        self.ss6 = IntVar()
        self.ss7 = IntVar()
        self.ss8 = IntVar()
        self.ss9 = IntVar()
        self.ss10 = IntVar()
        self.ss11 = IntVar()
        self.ss12 = IntVar()
        self.ss13 = IntVar()
        self.ss14 = IntVar()
        self.ss15 = IntVar()

        # =========== V section 3 ========
        self.sss1 = IntVar()
        self.sss2 = IntVar()
        self.sss3 = IntVar()
        self.sss4 = IntVar()
        self.sss5 = IntVar()
        self.sss6 = IntVar()
        self.sss7 = IntVar()
        self.sss8 = IntVar()
        self.sss9 = IntVar()
        self.sss10 = IntVar()
        self.sss11 = IntVar()
        self.sss12 = IntVar()
        self.sss13 = IntVar()
        self.sss14 = IntVar()
        self.sss15 = IntVar()

        # =========== V total =============
        self.section1 = StringVar()
        self.section2 = StringVar()
        self.section3 = StringVar()

        # ============custumer Data===============#
        F1 = Frame(root, bd=2, width=310, height=170, bg='#0B4C5f')
        F1.place(x=938, y=31)
        txt1 = Label(F1, text='CUSTOMER INFO', font=('tajawal', 13, 'bold'), fg='white', bg='#0B4C5f')
        txt1.place(x=100, y=0)
        cust_name = Label(F1, text='Name', font=('tajawal', 12), fg='white', bg='#0B4C5f')
        cust_name.place(x=5, y=40)
        cust_phone = Label(F1, text='Phone', font=('tajawal', 12), fg='white', bg='#0B4C5f')
        cust_phone.place(x=5, y=70)
        cust_bill = Label(F1, text='Bill Num', font=('tajawal', 12), fg='white', bg='#0B4C5f')
        cust_bill.place(x=5, y=100)

        En_name = Entry(F1, textvariable=self.name, justify='center')
        En_name.place(x=104, y=40)
        En_phone = Entry(F1, textvariable=self.phone, justify='center')
        En_phone.place(x=104, y=70)
        En_bill = Entry(F1, textvariable=self.bill, justify='center')
        En_bill.place(x=104, y=100)

        btn_cust = Button(F1, text='search', font=('tajawal', 10), width=17, height=1, bg='silver', fg='black')
        btn_cust.place(x=104, y=130)

        # ==============BILL=================#
        F2 = Frame(root, bd=2, width=309, height=320, bg='silver')
        F2.place(x=938, y=205)
        scrol_y = Scrollbar(F2, orient=VERTICAL)
        self.txtarea = Text(F2, yscrollcommand=scrol_y.set)
        scrol_y.pack(side=LEFT, fill=Y)
        scrol_y.config(command=self.txtarea.yview)
        self.txtarea.pack(fill=BOTH, expand=1)

        # ===============PRICE=================#
        F3 = Frame(root, bd=2, width=605, height=130, bg='#0B4C5f')
        F3.place(x=642, y=558)
        total = Button(F3, text='TOTAL', width=8, height=1, fg='black', bg='silver', font=('tajawal', 10),
                       command=self.total)
        total.place(x=500, y=5)
        bill = Button(F3, text='EXPORT', width=8, height=1, fg='black', bg='silver', font=('tajawal', 10),
                      command=self.billl)
        bill.place(x=500, y=44)
        clear = Button(F3, text='CLEAR', width=8, height=1, fg='black', bg='silver', font=('tajawal', 10),
                       command=self.clear)
        clear.place(x=500, y=83)

        txt2 = Label(F3, text='Section 1', font=('tajawal', 15, 'bold'), bg='#0B4C5f', fg='silver')
        txt2.place(x=5, y=10)
        txt3 = Label(F3, text='Section 2', font=('tajawal', 15, 'bold'), bg='#0B4C5f', fg='silver')
        txt3.place(x=5, y=50)
        txt4 = Label(F3, text='Section 3', font=('tajawal', 15, 'bold'), bg='#0B4C5f', fg='silver')
        txt4.place(x=5, y=90)

        En_1 = Entry(F3, textvariable=self.section1, width=24, justify='center')
        En_1.place(x=280, y=10)
        En_2 = Entry(F3, textvariable=self.section2, width=24, justify='center')
        En_2.place(x=280, y=50)
        En_3 = Entry(F3, textvariable=self.section3, width=24, justify='center')
        En_3.place(x=280, y=90)

        # ===================ITems[1]=====================
        F4 = Frame(root, width=318, height=660, bg='#0B4C5f')
        F4.place(x=1, y=31)
        text4 = Label(F4, text="( Juices and frozen )", font=('tajawal', 15, 'bold'), bg='#0B4C5f', fg='silver')
        text4.place(x=60, y=0)

        l1 = Label(F4, text='Burger', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l1.place(x=10, y=50)
        l2 = Label(F4, text='Mozzarella', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l2.place(x=10, y=95)
        l3 = Label(F4, text='Element 3', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l3.place(x=10, y=140)
        l4 = Label(F4, text='Element 4', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l4.place(x=10, y=178)
        l5 = Label(F4, text='Element 5', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l5.place(x=10, y=220)
        l6 = Label(F4, text='Element 6', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l6.place(x=10, y=260)
        l7 = Label(F4, text='Element 7', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l7.place(x=10, y=305)
        l8 = Label(F4, text='Element 8', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l8.place(x=10, y=345)
        l9 = Label(F4, text='Element 9', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l9.place(x=10, y=385)
        l10 = Label(F4, text='Element 1', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l10.place(x=10, y=425)
        l11 = Label(F4, text='Element 2', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l11.place(x=10, y=465)
        l12 = Label(F4, text='Element 3', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l12.place(x=10, y=505)
        l13 = Label(F4, text='Element 4', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l13.place(x=10, y=545)
        l14 = Label(F4, text='Element 6', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l14.place(x=10, y=585)
        l15 = Label(F4, text='Element 7', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        l15.place(x=10, y=625)

        # ( Button )
        lEnt1 = Entry(F4, textvariable=self.s1, width=12, justify='center')
        lEnt1.place(x=105, y=50)

        lBlus1 = Button(F4, text="+", width=1, bd=3, bg="silver", activebackground="silver", command=self.addS1)
        lBlus1.place(x=270, y=45)

        lMin1 = Button(F4, text="-", width=1, bd=3, bg="silver", activebackground="silver", command=self.subS1)
        lMin1.place(x=220, y=45)

        lEnt2 = Entry(F4, textvariable=self.s2, width=12, justify='center')
        lEnt2.place(x=105, y=95)

        lBlus2 = Button(F4, text="+", width=1, bd=3, bg="silver", command=self.addS2)
        lBlus2.place(x=270, y=90)

        lMin2 = Button(F4, text="-", width=1, bd=3, bg="silver", command=self.subS2)
        lMin2.place(x=220, y=90)

        lEnt3 = Entry(F4, textvariable=self.s3, width=12, justify='center')
        lEnt3.place(x=105, y=140)

        lBlus3 = Button(F4, text="+", width=1, bd=3, bg="silver", command=self.addS3)
        lBlus3.place(x=270, y=133)

        lMin3 = Button(F4, text="-", width=1, bd=3, bg="silver", command=self.subS3)
        lMin3.place(x=220, y=133)

        lEnt4 = Entry(F4, textvariable=self.s4, width=12, justify='center')
        lEnt4.place(x=105, y=178)

        lBlus4 = Button(F4, text="+", width=1, bd=3, bg="silver", command=self.addS4)
        lBlus4.place(x=270, y=175)

        lMin4 = Button(F4, text="-", width=1, bd=3, bg="silver", command=self.subS4)
        lMin4.place(x=220, y=175)

        lEnt5 = Entry(F4, textvariable=self.s5, width=12, justify='center')
        lEnt5.place(x=105, y=220)

        lBlus5 = Button(F4, text="+", width=1, bd=3, bg="silver", command=self.addS5)
        lBlus5.place(x=270, y=215)

        lMin5 = Button(F4, text="-", width=1, bd=3, bg="silver", command=self.subS5)
        lMin5.place(x=220, y=215)

        lEnt6 = Entry(F4, textvariable=self.s6, width=12, justify='center')
        lEnt6.place(x=105, y=260)

        lBlus6 = Button(F4, text="+", width=1, bd=3, bg="silver", command=self.addS6)
        lBlus6.place(x=270, y=255)

        lMin6 = Button(F4, text="-", width=1, bd=3, bg="silver", command=self.subS6)
        lMin6.place(x=220, y=255)

        lEnt7 = Entry(F4, textvariable=self.s7, width=12, justify='center')
        lEnt7.place(x=105, y=305)

        lBlus7 = Button(F4, text="+", width=1, bd=3, bg="silver", command=self.addS7)
        lBlus7.place(x=270, y=298)

        lMin7 = Button(F4, text="-", width=1, bd=3, bg="silver", command=self.subS7)
        lMin7.place(x=220, y=298)

        lEnt8 = Entry(F4, textvariable=self.s8, width=12, justify='center')
        lEnt8.place(x=105, y=345)

        lBlus8 = Button(F4, text="+", width=1, bd=3, bg="silver", command=self.addS8)
        lBlus8.place(x=270, y=340)

        lMin8 = Button(F4, text="-", width=1, bd=3, bg="silver", command=self.subS8)
        lMin8.place(x=220, y=340)

        lEnt9 = Entry(F4, textvariable=self.s9, width=12, justify='center')
        lEnt9.place(x=105, y=385)

        lBlus9 = Button(F4, text="+", width=1, bd=3, bg="silver", command=self.addS9)
        lBlus9.place(x=270, y=380)

        lMin9 = Button(F4, text="-", width=1, bd=3, bg="silver", command=self.subS9)
        lMin9.place(x=220, y=380)

        lEnt10 = Entry(F4, textvariable=self.s10, width=12, justify='center')
        lEnt10.place(x=105, y=425)

        lBlus10 = Button(F4, text="+", width=1, bd=3, bg="silver", command=self.addS10)
        lBlus10.place(x=270, y=420)

        lMin10 = Button(F4, text="-", width=1, bd=3, bg="silver", command=self.subS10)
        lMin10.place(x=220, y=420)

        lEnt11 = Entry(F4, textvariable=self.s11, width=12, justify='center')
        lEnt11.place(x=105, y=465)

        lBlus11 = Button(F4, text="+", width=1, bd=3, bg="silver", command=self.addS11)
        lBlus11.place(x=270, y=460)

        lMin11 = Button(F4, text="-", width=1, bd=3, bg="silver", command=self.subS11)
        lMin11.place(x=220, y=460)

        lEnt12 = Entry(F4, textvariable=self.s12, width=12, justify='center')
        lEnt12.place(x=105, y=505)

        lBlus12 = Button(F4, text="+", width=1, bd=3, bg="silver", command=self.addS12)
        lBlus12.place(x=270, y=500)

        lMin12 = Button(F4, text="-", width=1, bd=3, bg="silver", command=self.subS12)
        lMin12.place(x=220, y=500)

        lEnt13 = Entry(F4, textvariable=self.s13, width=12, justify='center')
        lEnt13.place(x=105, y=545)

        lBlus13 = Button(F4, text="+", width=1, bd=3, bg="silver", command=self.addS13)
        lBlus13.place(x=270, y=540)

        lMin13 = Button(F4, text="-", width=1, bd=3, bg="silver", command=self.subS13)
        lMin13.place(x=220, y=540)

        lEnt14 = Entry(F4, textvariable=self.s14, width=12, justify='center')
        lEnt14.place(x=105, y=585)

        lBlus14 = Button(F4, text="+", width=1, bd=3, bg="silver", command=self.addS14)
        lBlus14.place(x=270, y=580)

        lMin14 = Button(F4, text="-", width=1, bd=3, bg="silver", command=self.subS14)
        lMin14.place(x=220, y=580)

        lEnt15 = Entry(F4, textvariable=self.s15, width=12, justify='center')
        lEnt15.place(x=105, y=625)

        lBlus15 = Button(F4, text="+", width=1, bd=3, bg="silver", command=self.addS15)
        lBlus15.place(x=270, y=620)

        lMin15 = Button(F4, text="-", width=1, bd=3, bg="silver", command=self.subS15)
        lMin15.place(x=220, y=620)

        # ==========Items[2]=========================
        F5 = Frame(root, width=320, height=660, bg='#0B4C5f')
        F5.place(x=320, y=31)
        text5 = Label(F5, text='Section 2', font=('tajawal', 15, 'bold'), bg='#0B4C5f', fg='silver')
        text5.place(x=110, y=0)

        h1 = Label(F5, text='Element 1', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h1.place(x=10, y=50)
        h2 = Label(F5, text='Element 2', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h2.place(x=10, y=95)
        h3 = Label(F5, text='Element 3', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h3.place(x=10, y=140)
        h4 = Label(F5, text='Element 4', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h4.place(x=10, y=178)
        h5 = Label(F5, text='Element 5', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h5.place(x=10, y=220)
        h6 = Label(F5, text='Element 6', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h6.place(x=10, y=260)
        h7 = Label(F5, text='Element 7', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h7.place(x=10, y=305)
        h8 = Label(F5, text='Element 8', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h8.place(x=10, y=345)
        h9 = Label(F5, text='Element 9', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h9.place(x=10, y=385)
        h10 = Label(F5, text='Element 1', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h10.place(x=10, y=425)
        h11 = Label(F5, text='Element 2', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h11.place(x=10, y=465)
        h12 = Label(F5, text='Element 3', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h12.place(x=10, y=505)
        h13 = Label(F5, text='Element 4', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h13.place(x=10, y=545)
        h14 = Label(F5, text='Element 5', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h14.place(x=10, y=586)
        h15 = Label(F5, text='Element 6', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        h15.place(x=10, y=625)

        # Button
        hEnt1 = Entry(F5, textvariable=self.ss1, width=12, justify='center')
        hEnt1.place(x=105, y=50)

        hBlus1 = Button(F5, text="+", width=1, bd=3, bg="silver", activebackground="silver", command=self.addSS1)
        hBlus1.place(x=270, y=45)

        hMin1 = Button(F5, text="-", width=1, bd=3, bg="silver", activebackground="silver", command=self.subSS1)
        hMin1.place(x=220, y=45)

        hEnt2 = Entry(F5, textvariable=self.ss2, width=12, justify='center')
        hEnt2.place(x=105, y=95)

        hBlus2 = Button(F5, text="+", width=1, bd=3, bg="silver", command=self.addSS2)
        hBlus2.place(x=270, y=90)

        hMin2 = Button(F5, text="-", width=1, bd=3, bg="silver", command=self.subSS2)
        hMin2.place(x=220, y=90)

        hEnt3 = Entry(F5, textvariable=self.ss3, width=12, justify='center')
        hEnt3.place(x=105, y=140)

        hBlus3 = Button(F5, text="+", width=1, bd=3, bg="silver", command=self.addSS3)
        hBlus3.place(x=270, y=133)

        hMin3 = Button(F5, text="-", width=1, bd=3, bg="silver", command=self.subSS3)
        hMin3.place(x=220, y=133)

        hEnt4 = Entry(F5, textvariable=self.ss4, width=12, justify='center')
        hEnt4.place(x=105, y=178)

        hBlus4 = Button(F5, text="+", width=1, bd=3, bg="silver", command=self.addSS4)
        hBlus4.place(x=270, y=175)

        hMin4 = Button(F5, text="-", width=1, bd=3, bg="silver", command=self.subSS4)
        hMin4.place(x=220, y=175)

        hEnt5 = Entry(F5, textvariable=self.ss5, width=12, justify='center')
        hEnt5.place(x=105, y=220)

        hBlus5 = Button(F5, text="+", width=1, bd=3, bg="silver", command=self.addSS5)
        hBlus5.place(x=270, y=215)

        hMin5 = Button(F5, text="-", width=1, bd=3, bg="silver", command=self.subSS5)
        hMin5.place(x=220, y=215)

        hEnt6 = Entry(F5, textvariable=self.ss6, width=12, justify='center')
        hEnt6.place(x=105, y=260)

        hBlus6 = Button(F5, text="+", width=1, bd=3, bg="silver", command=self.addSS6)
        hBlus6.place(x=270, y=255)

        hMin6 = Button(F5, text="-", width=1, bd=3, bg="silver", command=self.subSS6)
        hMin6.place(x=220, y=255)

        hEnt7 = Entry(F5, textvariable=self.ss7, width=12, justify='center')
        hEnt7.place(x=105, y=305)

        hBlus7 = Button(F5, text="+", width=1, bd=3, bg="silver", command=self.addSS7)
        hBlus7.place(x=270, y=298)

        hMin7 = Button(F5, text="-", width=1, bd=3, bg="silver", command=self.subSS7)
        hMin7.place(x=220, y=298)

        hEnt8 = Entry(F5, textvariable=self.ss8, width=12, justify='center')
        hEnt8.place(x=105, y=345)

        hBlus8 = Button(F5, text="+", width=1, bd=3, bg="silver", command=self.addSS8)
        hBlus8.place(x=270, y=340)

        hMin8 = Button(F5, text="-", width=1, bd=3, bg="silver", command=self.subSS8)
        hMin8.place(x=220, y=340)

        hEnt9 = Entry(F5, textvariable=self.ss9, width=12, justify='center')
        hEnt9.place(x=105, y=385)

        hBlus9 = Button(F5, text="+", width=1, bd=3, bg="silver", command=self.addSS9)
        hBlus9.place(x=270, y=380)

        hMin9 = Button(F5, text="-", width=1, bd=3, bg="silver", command=self.subSS9)
        hMin9.place(x=220, y=380)

        hEnt10 = Entry(F5, textvariable=self.ss10, width=12, justify='center')
        hEnt10.place(x=105, y=425)

        hBlus10 = Button(F5, text="+", width=1, bd=3, bg="silver", command=self.addSS10)
        hBlus10.place(x=270, y=420)

        hMin10 = Button(F5, text="-", width=1, bd=3, bg="silver", command=self.subSS10)
        hMin10.place(x=220, y=420)

        hEnt11 = Entry(F5, textvariable=self.ss11, width=12, justify='center')
        hEnt11.place(x=105, y=465)

        hBlus11 = Button(F5, text="+", width=1, bd=3, bg="silver", command=self.addSS11)
        hBlus11.place(x=270, y=460)

        hMin11 = Button(F5, text="-", width=1, bd=3, bg="silver", command=self.subSS11)
        hMin11.place(x=220, y=460)

        hEnt12 = Entry(F5, textvariable=self.ss12, width=12, justify='center')
        hEnt12.place(x=105, y=505)

        hBlus12 = Button(F5, text="+", width=1, bd=3, bg="silver", command=self.addSS12)
        hBlus12.place(x=270, y=500)

        hMin12 = Button(F5, text="-", width=1, bd=3, bg="silver", command=self.subSS12)
        hMin12.place(x=220, y=500)

        hEnt13 = Entry(F5, textvariable=self.ss13, width=12, justify='center')
        hEnt13.place(x=105, y=545)

        hBlus13 = Button(F5, text="+", width=1, bd=3, bg="silver", command=self.addSS13)
        hBlus13.place(x=270, y=540)

        hMin13 = Button(F5, text="-", width=1, bd=3, bg="silver", command=self.subSS13)
        hMin13.place(x=220, y=540)

        hEnt14 = Entry(F5, textvariable=self.ss14, width=12, justify='center')
        hEnt14.place(x=105, y=585)

        hBlus14 = Button(F5, text="+", width=1, bd=3, bg="silver", command=self.addSS14)
        hBlus14.place(x=270, y=580)

        hMin14 = Button(F5, text="-", width=1, bd=3, bg="silver", command=self.subSS14)
        hMin14.place(x=220, y=580)

        hEnt15 = Entry(F5, textvariable=self.ss15, width=12, justify='center')
        hEnt15.place(x=105, y=625)

        hBlus15 = Button(F5, text="+", width=1, bd=3, bg="silver", command=self.addSS15)
        hBlus15.place(x=270, y=620)

        hMin15 = Button(F5, text="-", width=1, bd=3, bg="silver", command=self.subSS15)
        hMin15.place(x=220, y=620)

        # =============ITEMS[3]============================
        F6 = Frame(root, width=295, height=525, bg='#0B4C5f')
        F6.place(x=642, y=31)
        text6 = Label(F6, text='Section 3', font=('tajawal', 15, 'bold'), bg='#0B4C5f', fg='silver')
        text6.place(x=110, y=0)

        e1 = Label(F6, text='Element 1', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        e1.place(x=5, y=50)
        e2 = Label(F6, text='Element 2', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        e2.place(x=5, y=95)
        e3 = Label(F6, text='Element 3', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        e3.place(x=5, y=140)
        e4 = Label(F6, text='Element 4', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        e4.place(x=5, y=178)
        e5 = Label(F6, text='Element 5', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        e5.place(x=5, y=220)
        e6 = Label(F6, text='Element 6', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        e6.place(x=5, y=260)
        e7 = Label(F6, text='Element 7', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        e7.place(x=5, y=305)
        e8 = Label(F6, text='Element 8', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        e8.place(x=5, y=345)
        e9 = Label(F6, text='Element 9', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        e9.place(x=5, y=385)
        e10 = Label(F6, text='Element 1', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        e10.place(x=5, y=425)
        e11 = Label(F6, text='Element 2', font=('tajawal', 12), bg='#0B4C5f', fg='white')
        e11.place(x=5, y=465)

        # Button
        eEnt1 = Entry(F6, textvariable=self.sss1, width=12, justify='center')
        eEnt1.place(x=95, y=50)

        eBlus1 = Button(F6, text="+", width=1, bd=3, bg="silver", activebackground="silver", command=self.addSSS1)
        eBlus1.place(x=250, y=45)

        eMin1 = Button(F6, text="-", width=1, bd=3, bg="silver", activebackground="silver", command=self.subSSS1)
        eMin1.place(x=205, y=45)

        eEnt2 = Entry(F6, textvariable=self.sss2, width=12, justify='center')
        eEnt2.place(x=95, y=95)

        eBlus2 = Button(F6, text="+", width=1, bd=3, bg="silver", activebackground="silver", command=self.addSSS2)
        eBlus2.place(x=250, y=90)

        eMin2 = Button(F6, text="-", width=1, bd=3, bg="silver", activebackground="silver", command=self.subSSS2)
        eMin2.place(x=205, y=90)

        eEnt3 = Entry(F6, textvariable=self.sss3, width=12, justify='center')
        eEnt3.place(x=95, y=140)

        eBlus3 = Button(F6, text="+", width=1, bd=3, bg="silver", activebackground="silver", command=self.addSSS3)
        eBlus3.place(x=250, y=135)

        eMin3 = Button(F6, text="-", width=1, bd=3, bg="silver", activebackground="silver", command=self.subSSS3)
        eMin3.place(x=205, y=135)

        eEnt4 = Entry(F6, textvariable=self.sss4, width=12, justify='center')
        eEnt4.place(x=95, y=178)

        eBlus4 = Button(F6, text="+", width=1, bd=3, bg="silver", activebackground="silver", command=self.addSSS4)
        eBlus4.place(x=250, y=175)

        eMin4 = Button(F6, text="-", width=1, bd=3, bg="silver", activebackground="silver", command=self.subSSS4)
        eMin4.place(x=205, y=175)

        eEnt5 = Entry(F6, textvariable=self.sss5, width=12, justify='center')
        eEnt5.place(x=95, y=220)

        eBlus5 = Button(F6, text="+", width=1, bd=3, bg="silver", activebackground="silver", command=self.addSSS5)
        eBlus5.place(x=250, y=215)

        eMin5 = Button(F6, text="-", width=1, bd=3, bg="silver", activebackground="silver", command=self.subSSS5)
        eMin5.place(x=205, y=215)

        eEnt6 = Entry(F6, textvariable=self.sss6, width=12, justify='center')
        eEnt6.place(x=95, y=260)

        eBlus6 = Button(F6, text="+", width=1, bd=3, bg="silver", activebackground="silver", command=self.addSSS6)
        eBlus6.place(x=250, y=255)

        eMin6 = Button(F6, text="-", width=1, bd=3, bg="silver", activebackground="silver", command=self.subSSS6)
        eMin6.place(x=205, y=255)

        eEnt7 = Entry(F6, textvariable=self.sss7, width=12, justify='center')
        eEnt7.place(x=95, y=305)

        eBlus7 = Button(F6, text="+", width=1, bd=3, bg="silver", activebackground="silver", command=self.addSSS7)
        eBlus7.place(x=250, y=300)

        eMin7 = Button(F6, text="-", width=1, bd=3, bg="silver", activebackground="silver", command=self.subSSS7)
        eMin7.place(x=205, y=300)

        eEnt8 = Entry(F6, textvariable=self.sss8, width=12, justify='center')
        eEnt8.place(x=95, y=345)

        eBlus8 = Button(F6, text="+", width=1, bd=3, bg="silver", activebackground="silver", command=self.addSSS8)
        eBlus8.place(x=250, y=340)

        eMin8 = Button(F6, text="-", width=1, bd=3, bg="silver", activebackground="silver", command=self.subSSS8)
        eMin8.place(x=205, y=340)

        eEnt9 = Entry(F6, textvariable=self.sss9, width=12, justify='center')
        eEnt9.place(x=95, y=385)

        eBlus9 = Button(F6, text="+", width=1, bd=3, bg="silver", activebackground="silver", command=self.addSSS9)
        eBlus9.place(x=250, y=380)

        eMin9 = Button(F6, text="-", width=1, bd=3, bg="silver", activebackground="silver", command=self.subSSS9)
        eMin9.place(x=205, y=380)

        eEnt10 = Entry(F6, textvariable=self.sss10, width=12, justify='center')
        eEnt10.place(x=95, y=425)

        eBlus10 = Button(F6, text="+", width=1, bd=3, bg="silver", activebackground="silver", command=self.addSSS10)
        eBlus10.place(x=250, y=420)

        eMin10 = Button(F6, text="-", width=1, bd=3, bg="silver", activebackground="silver", command=self.subSSS10)
        eMin10.place(x=205, y=420)

        eEnt11 = Entry(F6, textvariable=self.sss11, width=12, justify='center')
        eEnt11.place(x=95, y=465)

        eBlus11 = Button(F6, text="+", width=1, bd=3, bg="silver", activebackground="silver", command=self.addSSS11)
        eBlus11.place(x=250, y=460)

        eMin11 = Button(F6, text="-", width=1, bd=3, bg="silver", activebackground="silver", command=self.subSSS11)
        eMin11.place(x=205, y=460)

        self.welcome()

    def total(self):
        self.ts1 = self.s1.get() * 1.5
        self.ts2 = self.s2.get() * 2
        self.ts3 = self.s3.get() * 1.5
        self.ts4 = self.s4.get() * 2
        self.ts5 = self.s5.get() * 1.5
        self.ts6 = self.s6.get() * 2
        self.ts7 = self.s7.get() * 1.5
        self.ts8 = self.s8.get() * 2
        self.ts9 = self.s9.get() * 1.5
        self.ts10 = self.s10.get() * 2
        self.ts11 = self.s11.get() * 1.5
        self.ts12 = self.s12.get() * 2
        self.ts13 = self.s13.get() * 1.5
        self.ts14 = self.s14.get() * 2
        self.ts15 = self.s15.get() * 1.5

        self.tsec1 = float(
            self.ts1 +
            self.ts2 +
            self.ts3 +
            self.ts4 +
            self.ts5 +
            self.ts6 +
            self.ts7 +
            self.ts8 +
            self.ts9 +
            self.ts10 +
            self.ts11 +
            self.ts12 +
            self.ts13 +
            self.ts14 +
            self.ts15
        )
        self.section1.set(str(self.tsec1) + " $ ")

        self.tss1 = self.ss1.get() * 1.5
        self.tss2 = self.ss2.get() * 2
        self.tss3 = self.ss3.get() * 1.5
        self.tss4 = self.ss4.get() * 2
        self.tss5 = self.ss5.get() * 1.5
        self.tss6 = self.ss6.get() * 2
        self.tss7 = self.ss7.get() * 1.5
        self.tss8 = self.ss8.get() * 2
        self.tss9 = self.ss9.get() * 1.5
        self.tss10 = self.ss10.get() * 2
        self.tss11 = self.ss11.get() * 1.5
        self.tss12 = self.ss12.get() * 2
        self.tss13 = self.ss13.get() * 1.5
        self.tss14 = self.ss14.get() * 2
        self.tss15 = self.ss15.get() * 1.5

        self.tsec2 = float(
            self.tss1 +
            self.tss2 +
            self.tss3 +
            self.tss4 +
            self.tss5 +
            self.tss6 +
            self.tss7 +
            self.tss8 +
            self.tss9 +
            self.tss10 +
            self.tss11 +
            self.tss12 +
            self.tss13 +
            self.tss14 +
            self.tss15
        )
        self.section2.set(str(self.tsec2) + " $ ")

        self.tsss1 = self.sss1.get() * 1.5
        self.tsss2 = self.sss2.get() * 2
        self.tsss3 = self.sss3.get() * 1.5
        self.tsss4 = self.sss4.get() * 2
        self.tsss5 = self.sss5.get() * 1.5
        self.tsss6 = self.sss6.get() * 2
        self.tsss7 = self.sss7.get() * 1.5
        self.tsss8 = self.sss8.get() * 2
        self.tsss9 = self.sss9.get() * 1.5
        self.tsss10 = self.sss10.get() * 2
        self.tsss11 = self.sss11.get() * 1.5
        self.tsec3 = float(
            self.tsss1 +
            self.tsss2 +
            self.tsss3 +
            self.tsss4 +
            self.tsss5 +
            self.tsss6 +
            self.tsss7 +
            self.tsss8 +
            self.tsss9 +
            self.tsss10 +
            self.tsss11
        )
        self.section3.set(str(self.tsec3) + " $ ")
        self.all = float(self.tsec1 + self.tsec2 + self.tsec3)

    def welcome(self):
        # self.txtarea.delete('1.0', END)
        self.txtarea.insert(END, "   Welcome in your SuperMarket")
        self.txtarea.insert(END, "\n===============================")
        self.txtarea.insert(END, f"\n\t B.NUM    : {self.bill.get()}")
        self.txtarea.insert(END, f"\n\t Name     : {self.name.get()}")
        self.txtarea.insert(END, f"\n\t PHONE    : {self.name.get()}")
        self.txtarea.insert(END, "\n===============================")
        self.txtarea.insert(END, f"\nPrice\t    Conter\t       Purchases")
        self.txtarea.insert(END, "\n==============================")

    def clear(self):
        self.s1.set(0)
        self.s2.set(0)
        self.s3.set(0)
        self.s4.set(0)
        self.s5.set(0)
        self.s6.set(0)
        self.s7.set(0)
        self.s8.set(0)
        self.s9.set(0)
        self.s10.set(0)
        self.s11.set(0)
        self.s12.set(0)
        self.s13.set(0)
        self.s14.set(0)
        self.s15.set(0)

        self.ss1.set(0)
        self.ss2.set(0)
        self.ss3.set(0)
        self.ss4.set(0)
        self.ss5.set(0)
        self.ss6.set(0)
        self.ss7.set(0)
        self.ss8.set(0)
        self.ss9.set(0)
        self.ss10.set(0)
        self.ss11.set(0)
        self.ss12.set(0)
        self.ss13.set(0)
        self.ss14.set(0)
        self.ss15.set(0)

        self.sss1.set(0)
        self.sss2.set(0)
        self.sss3.set(0)
        self.sss4.set(0)
        self.sss5.set(0)
        self.sss6.set(0)
        self.sss7.set(0)
        self.sss8.set(0)
        self.sss9.set(0)
        self.sss10.set(0)
        self.sss11.set(0)
        self.sss12.set(0)
        self.sss13.set(0)
        self.sss14.set(0)
        self.sss15.set(0)

        self.section1.set('')
        self.section2.set('')
        self.section3.set('')

        self.name.set('')
        self.phone.set('')
        self.bill.set('')

    def close(self):
        self.root.destroy()

    def save(self):
        op = messagebox.askyesno(" Save ", " Do you want save the bill ? ")
        if op > 0:
            self.bb = self.txtarea.get("1.0", END)
            f1 = open('majidsaqr' + str(self.bill.get()) + ".txt", "w", encoding="utf-8")
            f1.write(self.bb)
            f1.close()
        else:
            return

    def billl(self):

        #         self.welcome()
        if self.s1.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts1}\t\t{self.s1.get()}\t\tElement1 ")

        if self.s2.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts2}\t\t{self.s2.get()}\t\tElement2 ")

        if self.s3.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts3}\t\t{self.s3.get()}\t\tElement3 ")

        if self.s4.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts4}\t\t{self.s4.get()}\t\tElement4 ")

        if self.s5.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts5}\t\t{self.s5.get()}\t\tElement5 ")

        if self.s6.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts6}\t\t{self.s6.get()}\t\tElement6 ")

        if self.s7.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts7}\t\t{self.s7.get()}\t\tElement7 ")

        if self.s8.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts8}\t\t{self.s8.get()}\t\tElement8 ")

        if self.s9.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts9}\t\t{self.s9.get()}\t\tElement9 ")

        if self.s10.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts10}\t\t{self.s10.get()}\t\tElement10 ")

        if self.s11.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts11}\t\t{self.s11.get()}\t\tElement11 ")

        if self.s12.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts12}\t\t{self.s12.get()}\t\tElement12 ")

        if self.s13.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts13}\t\t{self.s13.get()}\t\tElement13 ")

        if self.s14.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts14}\t\t{self.s14.get()}\t\tElement14 ")

        if self.s15.get() != 0:
            self.txtarea.insert(END, f"\n {self.ts15}\t\t{self.s15.get()}\t\tElement15 ")

        if self.ss1.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss1}\t\t{self.ss1.get()}\t\tElement1 ")

        if self.ss2.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss2}\t\t{self.ss2.get()}\t\tElement2 ")

        if self.ss3.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss3}\t\t{self.ss3.get()}\t\tElement3 ")

        if self.ss4.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss4}\t\t{self.ss4.get()}\t\tElement4 ")

        if self.ss5.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss5}\t\t{self.ss5.get()}\t\tElement5 ")

        if self.ss5.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss5}\t\t{self.ss5.get()}\t\tElement5 ")

        if self.ss6.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss6}\t\t{self.ss6.get()}\t\tElement6 ")

        if self.ss7.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss7}\t\t{self.ss7.get()}\t\tElement7 ")

        if self.ss8.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss8}\t\t{self.ss8.get()}\t\tElement8 ")

        if self.ss9.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss9}\t\t{self.ss9.get()}\t\tElement9 ")

        if self.ss10.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss10}\t\t{self.ss10.get()}\t\tElement10 ")

        if self.ss11.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss11}\t\t{self.ss11.get()}\t\tElement11 ")

        if self.ss12.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss12}\t\t{self.ss12.get()}\t\tElement12 ")

        if self.ss13.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss13}\t\t{self.ss13.get()}\t\tElement13 ")

        if self.ss14.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss14}\t\t{self.ss14.get()}\t\tElement14 ")

        if self.ss15.get() != 0:
            self.txtarea.insert(END, f"\n {self.tss15}\t\t{self.ss15.get()}\t\tElement15 ")

        if self.sss1.get() != 0:
            self.txtarea.insert(END, f"\n {self.tsss1}\t\t{self.sss1.get()}\t\tElement1 ")

        if self.sss2.get() != 0:
            self.txtarea.insert(END, f"\n {self.tsss2}\t\t{self.sss2.get()}\t\tElement2 ")

        if self.sss3.get() != 0:
            self.txtarea.insert(END, f"\n {self.tsss3}\t\t{self.sss3.get()}\t\tElement3 ")

        if self.sss4.get() != 0:
            self.txtarea.insert(END, f"\n {self.tsss4}\t\t{self.sss4.get()}\t\tElement4 ")

        if self.sss5.get() != 0:
            self.txtarea.insert(END, f"\n {self.tsss5}\t\t{self.sss5.get()}\t\tElement5 ")

        if self.sss6.get() != 0:
            self.txtarea.insert(END, f"\n {self.tsss6}\t\t{self.sss6.get()}\t\tElement6 ")

        if self.sss7.get() != 0:
            self.txtarea.insert(END, f"\n {self.tsss7}\t\t{self.sss7.get()}\t\tElement7 ")

        if self.sss8.get() != 0:
            self.txtarea.insert(END, f"\n {self.tsss8}\t\t{self.sss8.get()}\t\tElement8 ")

        if self.sss9.get() != 0:
            self.txtarea.insert(END, f"\n {self.tsss9}\t\t{self.sss9.get()}\t\tElement9 ")

        if self.sss10.get() != 0:
            self.txtarea.insert(END, f"\n {self.tsss10}\t\t{self.sss10.get()}\t\tElement10 ")

        if self.sss11.get() != 0:
            self.txtarea.insert(END, f"\n {self.tsss11}\t\t{self.sss11.get()}\t\tElement11 ")

        self.txtarea.insert(END, "\n.......................................")
        self.txtarea.insert(END, f"\nTotal Price $\t             {self.all}")
        self.txtarea.insert(END, "\n.......................................")
        self.save()

    # ( FUNCTIONS ADD & MIN SECTION 1)
    def addS1(self):
        value = self.s1.get()
        value = value + 1
        self.s1.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A1'] = value
        wb.save("sample_file.xlsx")

    def subS1(self):
        value = self.s1.get()
        if value > 0:
            value = value - 1
        self.s1.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A1'] = value
        wb.save("sample_file.xlsx")

    def addS2(self):
        value = self.s2.get()
        value = value + 1
        self.s2.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A2'] = value
        wb.save("sample_file.xlsx")

    def subS2(self):
        value = self.s2.get()
        if value > 0:
            value = value - 1
        self.s2.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A2'] = value
        wb.save("sample_file.xlsx")

    def addS3(self):
        value = self.s3.get()
        value = value + 1
        self.s3.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A3'] = value
        wb.save("sample_file.xlsx")

    def subS3(self):
        value = self.s3.get()
        if value > 0:
            value = value - 1
        self.s3.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A3'] = value
        wb.save("sample_file.xlsx")

    def addS4(self):
        value = self.s4.get()
        value = value + 1
        self.s4.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A4'] = value
        wb.save("sample_file.xlsx")

    def subS4(self):
        value = self.s4.get()
        if value > 0:
            value = value - 1
        self.s4.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A4'] = value
        wb.save("sample_file.xlsx")

    def addS5(self):
        value = self.s5.get()
        value = value + 1
        self.s5.set(value)

    def subS5(self):
        value = self.s5.get()
        if value > 0:
            value = value - 1
        self.s5.set(value)

    def addS6(self):
        value = self.s6.get()
        value = value + 1
        self.s6.set(value)

    def subS6(self):
        value = self.s6.get()
        if value > 0:
            value = value - 1
        self.s6.set(value)

    def addS7(self):
        value = self.s7.get()
        value = value + 1
        self.s7.set(value)

    def subS7(self):
        value = self.s7.get()
        if value > 0:
            value = value - 1
        self.s7.set(value)

    def addS8(self):
        value = self.s8.get()
        value = value + 1
        self.s8.set(value)

    def subS8(self):
        value = self.s8.get()
        if value > 0:
            value = value - 1
        self.s8.set(value)

    def addS9(self):
        value = self.s9.get()
        value = value + 1
        self.s9.set(value)

    def subS9(self):
        value = self.s9.get()
        if value > 0:
            value = value - 1
        self.s9.set(value)

    def addS10(self):
        value = self.s10.get()
        value = value + 1
        self.s10.set(value)

    def subS10(self):
        value = self.s10.get()
        if value > 0:
            value = value - 1
        self.s10.set(value)

    def addS11(self):
        value = self.s11.get()
        value = value + 1
        self.s11.set(value)

    def subS11(self):
        value = self.s11.get()
        if value > 0:
            value = value - 1
        self.s11.set(value)

    def addS12(self):
        value = self.s12.get()
        value = value + 1
        self.s12.set(value)

    def subS12(self):
        value = self.s12.get()
        if value > 0:
            value = value - 1
        self.s12.set(value)

    def addS13(self):
        value = self.s13.get()
        value = value + 1
        self.s13.set(value)

    def subS13(self):
        value = self.s13.get()
        if value > 0:
            value = value - 1
        self.s13.set(value)

    def addS14(self):
        value = self.s14.get()
        value = value + 1
        self.s14.set(value)

    def subS14(self):
        value = self.s14.get()
        if value > 0:
            value = value - 1
        self.s14.set(value)

    def addS15(self):
        value = self.s15.get()
        value = value + 1
        self.s15.set(value)

    def subS15(self):
        value = self.s15.get()
        if value > 0:
            value = value - 1
        self.s15.set(value)

    # ( FUNCTIONS ADD & MIN SECTION 2 )

    def addSS1(self):
        value = self.ss1.get()
        value = value + 1
        self.ss1.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A1'] = value
        wb.save("sample_file.xlsx")

    def subSS1(self):
        value = self.ss1.get()
        if value > 0:
            value = value - 1
        self.ss1.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A1'] = value
        wb.save("sample_file.xlsx")

    def addSS2(self):
        value = self.ss2.get()
        value = value + 1
        self.ss2.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A2'] = value
        wb.save("sample_file.xlsx")

    def subSS2(self):
        value = self.ss2.get()
        if value > 0:
            value = value - 1
        self.ss2.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A2'] = value
        wb.save("sample_file.xlsx")

    def addSS3(self):
        value = self.ss3.get()
        value = value + 1
        self.ss3.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A3'] = value
        wb.save("sample_file.xlsx")

    def subSS3(self):
        value = self.ss3.get()
        if value > 0:
            value = value - 1
        self.ss3.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A3'] = value
        wb.save("sample_file.xlsx")

    def addSS4(self):
        value = self.ss4.get()
        value = value + 1
        self.ss4.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A4'] = value
        wb.save("sample_file.xlsx")

    def subSS4(self):
        value = self.ss4.get()
        if value > 0:
            value = value - 1
        self.ss4.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A4'] = value
        wb.save("sample_file.xlsx")

    def addSS5(self):
        value = self.ss5.get()
        value = value + 1
        self.ss5.set(value)

    def subSS5(self):
        value = self.ss5.get()
        if value > 0:
            value = value - 1
        self.ss5.set(value)

    def addSS6(self):
        value = self.ss6.get()
        value = value + 1
        self.ss6.set(value)

    def subSS6(self):
        value = self.ss6.get()
        if value > 0:
            value = value - 1
        self.ss6.set(value)

    def addSS7(self):
        value = self.ss7.get()
        value = value + 1
        self.ss7.set(value)

    def subSS7(self):
        value = self.ss7.get()
        if value > 0:
            value = value - 1
        self.ss7.set(value)

    def addSS8(self):
        value = self.ss8.get()
        value = value + 1
        self.ss8.set(value)

    def subSS8(self):
        value = self.ss8.get()
        if value > 0:
            value = value - 1
        self.ss8.set(value)

    def addSS9(self):
        value = self.ss9.get()
        value = value + 1
        self.ss9.set(value)

    def subSS9(self):
        value = self.ss9.get()
        if value > 0:
            value = value - 1
        self.ss9.set(value)

    def addSS10(self):
        value = self.ss10.get()
        value = value + 1
        self.ss10.set(value)

    def subSS10(self):
        value = self.ss10.get()
        if value > 0:
            value = value - 1
        self.ss10.set(value)

    def addSS11(self):
        value = self.ss11.get()
        value = value + 1
        self.ss11.set(value)

    def subSS11(self):
        value = self.ss11.get()
        if value > 0:
            value = value - 1
        self.ss11.set(value)

    def addSS12(self):
        value = self.ss12.get()
        value = value + 1
        self.ss12.set(value)

    def subSS12(self):
        value = self.ss12.get()
        if value > 0:
            value = value - 1
        self.ss12.set(value)

    def addSS13(self):
        value = self.ss13.get()
        value = value + 1
        self.ss13.set(value)

    def subSS13(self):
        value = self.ss13.get()
        if value > 0:
            value = value - 1
        self.ss13.set(value)

    def addSS14(self):
        value = self.ss14.get()
        value = value + 1
        self.ss14.set(value)

    def subSS14(self):
        value = self.ss14.get()
        if value > 0:
            value = value - 1
        self.ss14.set(value)

    def addSS15(self):
        value = self.ss15.get()
        value = value + 1
        self.ss15.set(value)

    def subSS15(self):
        value = self.s15.get()
        if value > 0:
            value = value - 1
        self.ss15.set(value)

    # ( FUNCTIONS ADD & MIN SECTION 3 )

    def addSSS1(self):
        value = self.sss1.get()
        value = value + 1
        self.sss1.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A1'] = value
        wb.save("sample_file.xlsx")

    def subSSS1(self):
        value = self.sss1.get()
        if value > 0:
            value = value - 1
        self.sss1.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A1'] = value
        wb.save("sample_file.xlsx")

    def addSSS2(self):
        value = self.sss2.get()
        value = value + 1
        self.sss2.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A2'] = value
        wb.save("sample_file.xlsx")

    def subSSS2(self):
        value = self.sss2.get()
        if value > 0:
            value = value - 1
        self.sss2.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A2'] = value
        wb.save("sample_file.xlsx")

    def addSSS3(self):
        value = self.sss3.get()
        value = value + 1
        self.sss3.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A3'] = value
        wb.save("sample_file.xlsx")

    def subSSS3(self):
        value = self.sss3.get()
        if value > 0:
            value = value - 1
        self.sss3.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A3'] = value
        wb.save("sample_file.xlsx")

    def addSSS4(self):
        value = self.sss4.get()
        value = value + 1
        self.sss4.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A4'] = value
        wb.save("sample_file.xlsx")

    def subSSS4(self):
        value = self.sss4.get()
        if value > 0:
            value = value - 1
        self.sss4.set(value)
        wb = load_workbook("sample_file.xlsx")
        sheet = wb.active

        sheet['A4'] = value
        wb.save("sample_file.xlsx")

    def addSSS5(self):
        value = self.sss5.get()
        value = value + 1
        self.sss5.set(value)

    def subSSS5(self):
        value = self.sss5.get()
        if value > 0:
            value = value - 1
        self.sss5.set(value)

    def addSSS6(self):
        value = self.sss6.get()
        value = value + 1
        self.sss6.set(value)

    def subSSS6(self):
        value = self.sss6.get()
        if value > 0:
            value = value - 1
        self.sss6.set(value)

    def addSSS7(self):
        value = self.sss7.get()
        value = value + 1
        self.sss7.set(value)

    def subSSS7(self):
        value = self.sss7.get()
        if value > 0:
            value = value - 1
        self.sss7.set(value)

    def addSSS8(self):
        value = self.sss8.get()
        value = value + 1
        self.sss8.set(value)

    def subSSS8(self):
        value = self.sss8.get()
        if value > 0:
            value = value - 1
        self.sss8.set(value)

    def addSSS9(self):
        value = self.sss9.get()
        value = value + 1
        self.sss9.set(value)

    def subSSS9(self):
        value = self.sss9.get()
        if value > 0:
            value = value - 1
        self.sss9.set(value)

    def addSSS10(self):
        value = self.sss10.get()
        value = value + 1
        self.sss10.set(value)

    def subSSS10(self):
        value = self.sss10.get()
        if value > 0:
            value = value - 1
        self.sss10.set(value)

    def addSSS11(self):
        value = self.sss11.get()
        value = value + 1
        self.sss11.set(value)

    def subSSS11(self):
        value = self.sss11.get()
        if value > 0:
            value = value - 1
        self.sss11.set(value)


root = Tk()
window = supermarket(root)
root.mainloop()
